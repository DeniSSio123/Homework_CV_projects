import numpy as np
import psycopg2
from PyQt5 import QtCore, QtGui, QtWidgets, QtSql
from openpyxl import load_workbook

db = QtSql.QSqlDatabase.addDatabase('QPSQL')


# Модуль QtCore содержит классы не связанные с реализацией графического интерфейса. От этого модуля зависят все остальные модули PyQt.
# Модуль QtGui содержит классы, реализующие комnоненты графического интерфейса(надписи, кнопки, текстовые поля и др.)
# Модуль QtSql- включает поддержку баз данных SQLite, MySQL и др.;

class LoginPage(QtWidgets.QWidget):  # создает класс, который наследуется от класса QWidget
    def Pressed_OK(self):
        haser = QtCore.QCryptographicHash(QtCore.QCryptographicHash.Sha3_384)
        haser.addData(self.password_edit.text().encode())
        password = str(haser.result().toHex(), 'UTF-8')
        global db
        qw = QtCore.QUuid.createUuid()
        qw.toString()
        db.setHostName("localhost")
        db.setPort(5432)
        db.setDatabaseName("data")
        db.setUserName(self.login_edit.text())
        db.setPassword(password)
        if db.open() == True:
            db.close()
            self.login_edit.clear()
            self.password_edit.clear()
            men = Menu()
            men.BackBottom.clicked.connect(lambda: self.stack.setCurrentIndex(0))
            self.stack.addWidget(men)
            self.stack.setCurrentIndex(1)
        else:
            db.close()
            mess = QtWidgets.QMessageBox(self)
            mess.setText("Неправильно введен логин или пароль")
            mess.exec()
            self.password_edit.clear()

    def __init__(self):  # создает конструкцию класса
        super().__init__()  # возвращает родительский объект LoginPage с классом
        self.resize(1000, 1000)  # минимальные размеры окна
        font = QtGui.QFont()  # метод для создания GUI
        font.setPointSize(14)
        self.LogWin = QtWidgets.QWidget()  # создаёт объект окна с помощью класса QWidget

        self.login_label = QtWidgets.QLabel('Login', self.LogWin)
        self.login_label.setGeometry(
            QtCore.QRect(470, 200, 150, 70))  # Класс QRect оnисывает координаты и размеры прямоугольной области.
        self.login_label.setFont(font)

        self.password_label = QtWidgets.QLabel('Password', self.LogWin)
        self.password_label.setGeometry(QtCore.QRect(450, 400, 150, 70))
        self.password_label.setFont(font)

        self.login_edit = QtWidgets.QLineEdit(self.LogWin)
        self.login_edit.setGeometry(QtCore.QRect(300, 270, 400, 70))
        self.login_edit.setFont(font)
        self.login_edit.setToolTip('Введите ваш логин')

        self.password_edit = QtWidgets.QLineEdit(self.LogWin)
        self.password_edit.setGeometry(QtCore.QRect(300, 470, 400, 70))
        self.password_edit.setEchoMode(QtWidgets.QLineEdit.Password)
        self.password_edit.setFont(font)
        self.password_edit.setToolTip('Пароль должен быть не менее 6-ти символов.\n'
                                      'Пароль должен содержать минимум 1 цифру и 1 букву.')

        self.OK_buttom = QtWidgets.QPushButton(self.LogWin)
        self.OK_buttom.setGeometry((QtCore.QRect(460, 600, 100, 90)))
        self.OK_buttom.setText('OK')

        self.stack = QtWidgets.QStackedLayout(self)
        self.stack.addWidget(self.LogWin)

        self.OK_buttom.clicked.connect(lambda: self.Pressed_OK())
        self.stack.setCurrentIndex(0)


class Menu(QtWidgets.QWidget):
    def BackBottomAll(self):
        self.stack.removeWidget(self.stack.currentWidget())
        self.stack.setCurrentIndex(0)

    def CreateSubWindow(self, cl):
        obj = cl()
        obj.BackButtom.clicked.connect(lambda: self.BackBottomAll())
        self.stack.addWidget(obj)
        self.stack.setCurrentIndex(1)

    def __init__(self):
        super().__init__()  # возвращает родительский объект Menu с классом
        self.resize(900, 1000)  # минимальные размеры окна
        font = QtGui.QFont()  # метод для создания GUI
        font.setPointSize(14)
        self.win = QtWidgets.QWidget()  # создаёт объект окна с помощью класса QWidget

        self.cre_plan = QtWidgets.QPushButton(self.win)  # создаем виджет кнопки
        self.cre_plan.setGeometry(QtCore.QRect(300, 200, 400, 100))
        self.cre_plan.setText('Создать план')
        self.cre_plan.setFont(font)

        self.chan_plan = QtWidgets.QPushButton(self.win)
        self.chan_plan.setGeometry(QtCore.QRect(300, 400, 400, 100))
        self.chan_plan.setText('Изменить план')
        self.chan_plan.setFont(font)

        self.upload_plan = QtWidgets.QPushButton(self.win)
        self.upload_plan.setGeometry(QtCore.QRect(300, 600, 400, 100))
        self.upload_plan.setText('Загрузить план')
        self.upload_plan.setFont(font)

        self.BackBottom = QtWidgets.QPushButton('Back', self.win)
        self.BackBottom.setGeometry(700, 750, 100, 50)

        self.stack = QtWidgets.QStackedLayout(self)
        self.stack.addWidget(self.win)

        self.cre_plan.clicked.connect(lambda: self.CreateSubWindow(CreatePlanPosition))
        self.chan_plan.clicked.connect(lambda: self.CreateSubWindow(ChangePlan))
        self.upload_plan.clicked.connect(lambda: self.CreateSubWindow(UploadPlan))
        self.stack.setCurrentIndex(0)


class LineEditCompleter(QtWidgets.QLineEdit):
    def check(self):
        if self.text().strip() not in self.data and self.text() != '':
            messageBox = QtWidgets.QMessageBox(self.parent())
            messageBox.setText("Неправильно введено значение, оно будет очищенно")
            erPalette = QtGui.QPalette()
            erPalette.setColor(QtGui.QPalette.Base, QtGui.QColor(255, 204, 204))
            self.setPalette(erPalette)
            messageBox.exec()
            Palette = QtGui.QPalette()
            Palette.setColor(QtGui.QPalette.Base, QtGui.QColor(255, 255, 255, 255))
            self.setPalette(Palette)
            self.clear()

    def __init__(self, parent, db, query):
        super().__init__(parent)
        self.model = QtSql.QSqlTableModel(db=db)
        quer = QtSql.QSqlQuery(query, db=db)
        self.model.setQuery(quer)
        self.complet = QtWidgets.QCompleter()
        self.complet.setModel(self.model)
        self.complet.setCompletionColumn(0)
        self.setCompleter(self.complet)
        self.setClearButtonEnabled(True)
        self.editingFinished.connect(lambda: self.check())
        self.data = [self.completer().model().record(i).value(0) for i in range(0, self.completer().model().rowCount())]


class ComboBoxCompleter(QtWidgets.QComboBox):
    def __init__(self, parent, db, query):
        super().__init__(parent)
        self.model = QtSql.QSqlTableModel(db=db)
        quer = QtSql.QSqlQuery(query, db=db)
        self.model.setQuery(quer)
        self.setModel(self.model)
        self.setModelColumn(0)
        self.setCurrentIndex(-1)


class CalendarBox(QtWidgets.QDialog):
    def __init__(self, parent, label_1, label_2, start_stop):
        super().__init__(parent)
        self.cal = QtWidgets.QCalendarWidget(self)
        self.cal.setGridVisible(True)
        if (start_stop == 'start') and ('*' not in label_2.text()):
            self.cal.setMaximumDate(
                QtCore.QDate(int(label_2.text()[6:]), int(label_2.text()[3:5]), int(label_2.text()[:2])))
        elif (start_stop == 'stop') and ('*' not in label_2.text()):
            self.cal.setMinimumDate(
                QtCore.QDate(int(label_2.text()[6:]), int(label_2.text()[3:5]), int(label_2.text()[:2])))
        self.cal.move(10, 10)
        self.cal.setGeometry(QtCore.QRect(5, 5, 464, 289))
        self.cal.clicked[QtCore.QDate].connect(self.showDate)
        self.date = self.cal.selectedDate()

        self.lbl = QtWidgets.QLabel(self)
        self.lbl.setText(self.date.toString("dd.MM.yyyy"))
        self.lbl.setGeometry(187, 300, 100, 30)

        self.okb = QtWidgets.QPushButton(self)
        self.okb.setGeometry(QtCore.QRect(217, 330, 40, 40))
        self.okb.setText("ok")
        self.okb.clicked.connect(lambda: self.press(label_1))

        self.setWindowTitle('Calendar')
        self.setBackgroundRole(QtGui.QPalette.Window)
        self.setVisible(True)

    def showDate(self, date):
        self.lbl.setText(date.toString("dd.MM.yyyy"))

    def press(self, label_1):
        label_1.setText(self.lbl.text())
        self.destroy()


class CreateF2(QtWidgets.QWidget):
    def calen(self, label_1, label_2, start_stop):
        cal = CalendarBox(self, label_1, label_2, start_stop)
        cal.resize(474, 373)
        cal.show()

    def __init__(self):
        super().__init__()  # возвращает родительский объект CreatePlan с классом
        # self.resize(1000, 1300)  # минимальные размеры окна
        font = QtGui.QFont()  # метод для создания GUI
        font.setPointSize(10)
        global db
        db.setHostName("localhost")
        db.setPort(5432)
        db.setDatabaseName("data")
        db.setUserName('browser')
        db.setPassword("browser")
        db.open()

        calendarPixMap = QtGui.QPixmap("C:/Users/Валерия/PyCharm/GovBuy/design/icons/img_52517.png")
        calendarButIcon = QtGui.QIcon(calendarPixMap)

        self.pn_ext_ident = QtWidgets.QLabel('Внешний идентификатор позиции', self)
        self.pn_ext_ident.setGeometry(QtCore.QRect(70, 0, 500, 40))
        self.pn_ext_ident.setFont(font)
        self.line_pn_ext_ident = QtWidgets.QLineEdit(self)
        self.line_pn_ext_ident.setGeometry(QtCore.QRect(580, 10, 400, 30))
        self.line_pn_ext_ident.setFont(font)
        self.line_pn_ext_ident.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("[1-9]\d+(\.|,)\d+")))
        self.line_pn_ext_ident.setClearButtonEnabled(1)
        self.line_pn_ext_ident.setToolTip('Начинайте вводить внешний идентификатор позиции')

        self.ContractCurrency = QtWidgets.QLabel('Валюта договора', self)
        self.ContractCurrency.setGeometry(QtCore.QRect(70, 45, 500, 40))
        self.ContractCurrency.setFont(font)
        self.lineContractCurrency = ComboBoxCompleter(self, db, "select v_name from nm_currencies")
        self.lineContractCurrency.setGeometry(QtCore.QRect(580, 46, 400, 40))
        self.lineContractCurrency.setFont(font)

        self.PaymentAmount = QtWidgets.QLabel('Объем оплаты долгосрочного договора', self)  # 4 digits
        self.PaymentAmount.setGeometry(QtCore.QRect(70, 95, 500, 40))
        self.PaymentAmount.setFont(font)
        self.linePaymentAmount = QtWidgets.QLineEdit(self)
        self.linePaymentAmount.setGeometry(QtCore.QRect(580, 97, 400, 40))
        self.linePaymentAmount.setFont(font)
        self.linePaymentAmount.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("\d+(\.|,)\d+")))
        self.linePaymentAmount.setClearButtonEnabled(1)

        self.TransCurrency = QtWidgets.QLabel('Валюта объема оплаты долгосрочного договора', self)
        self.TransCurrency.setGeometry(QtCore.QRect(70, 145, 500, 40))
        self.TransCurrency.setFont(font)
        self.lineTransCurrency = ComboBoxCompleter(self, db, "select v_name from nm_currencies")
        self.lineTransCurrency.setGeometry(QtCore.QRect(580, 150, 400, 40))
        self.lineTransCurrency.setFont(font)

        self.AttractionValue = QtWidgets.QLabel('Объем привлечения субъектов малого и среднего\nпредпринимательства',
                                                self)
        self.AttractionValue.setGeometry(QtCore.QRect(70, 200, 500, 50))
        self.AttractionValue.setFont(font)
        self.lineAttractionValue = QtWidgets.QLineEdit(self)
        self.lineAttractionValue.setGeometry(QtCore.QRect(580, 202, 400, 40))
        self.lineAttractionValue.setFont(font)
        self.lineAttractionValue.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("\d+(\.|,)\d+")))
        self.lineAttractionValue.setClearButtonEnabled(1)

        self.Value = QtWidgets.QLabel('Количество (объем)', self)
        self.Value.setGeometry(QtCore.QRect(70, 265, 500, 40))
        self.Value.setFont(font)
        self.lineValue = QtWidgets.QLineEdit(self)
        self.lineValue.setGeometry(QtCore.QRect(580, 260, 400, 40))
        self.lineValue.setFont(font)
        self.lineValue.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("\d+(\.|,)\d+")))
        self.lineValue.setClearButtonEnabled(1)

        self.MeasureUnit = QtWidgets.QLabel('Единица измерения', self)
        self.MeasureUnit.setGeometry(QtCore.QRect(70, 310, 500, 50))
        self.MeasureUnit.setFont(font)
        self.lineMeasureUnit = LineEditCompleter(self, db, "select v_name from nm_measure_units;")
        self.lineMeasureUnit.setGeometry(QtCore.QRect(580, 315, 400, 40))
        self.lineMeasureUnit.setFont(font)

        self.SmallBusiness = QtWidgets.QLabel(
            'Закупка, участниками которой являются только\nсубъекты малого и среднего предпринимательства', self)
        self.SmallBusiness.setGeometry(QtCore.QRect(70, 370, 550, 50))
        self.SmallBusiness.setFont(font)
        self.SmallBusinessButGroup = QtWidgets.QButtonGroup(self)
        SmallBusinessbuttonYes = QtWidgets.QRadioButton(self)
        SmallBusinessbuttonYes.setGeometry((QtCore.QRect(590, 375, 100, 50)))
        SmallBusinessbuttonYes.setText("Да")
        self.SmallBusinessButGroup.addButton(SmallBusinessbuttonYes)
        SmallBusinessbuttonNo = QtWidgets.QRadioButton(self)
        SmallBusinessbuttonNo.setGeometry((QtCore.QRect(730, 375, 100, 50)))
        SmallBusinessbuttonNo.setText("Нет")
        self.SmallBusinessButGroup.addButton(SmallBusinessbuttonNo)

        self.Exclude = QtWidgets.QLabel(
            'Закупка исключается при расчете годового объема\nзакупок, участниками которых являются субъекты\nмалого и среднего предпринимательства',
            self)
        self.Exclude.setGeometry(QtCore.QRect(70, 438, 550, 70))
        self.Exclude.setFont(font)
        self.ExcludeButGroup = QtWidgets.QButtonGroup(self)
        ExcludeRadiobuttonYes = QtWidgets.QRadioButton(self)
        ExcludeRadiobuttonYes.setText("Да")
        ExcludeRadiobuttonYes.setGeometry(QtCore.QRect(590, 452, 100, 50))
        self.ExcludeButGroup.addButton(ExcludeRadiobuttonYes)
        ExcludeRadiobuttonNo = QtWidgets.QRadioButton(self)
        ExcludeRadiobuttonNo.setText("Нет")
        ExcludeRadiobuttonNo.setGeometry((QtCore.QRect(730, 452, 100, 50)))
        self.ExcludeButGroup.addButton(ExcludeRadiobuttonNo)

        self.Innovation = QtWidgets.QLabel(
            'Закупка товаров, работ, услуг, удовлетворяющих\nкритериям отнесения кинновационной продукции,\nвысокотехнологичной продукции',
            self)
        self.Innovation.setGeometry(QtCore.QRect(70, 530, 550, 70))
        self.Innovation.setFont(font)
        self.InnovationButGroup = QtWidgets.QButtonGroup(self)
        InovationRadiobuttonYes = QtWidgets.QRadioButton(self)
        InovationRadiobuttonYes.setGeometry((QtCore.QRect(590, 530, 100, 50)))
        InovationRadiobuttonYes.setText("Да")
        self.InnovationButGroup.addButton(InovationRadiobuttonYes)
        InovationRadiobuttonNo = QtWidgets.QRadioButton(self)
        InovationRadiobuttonNo.setGeometry((QtCore.QRect(730, 530, 100, 50)))
        InovationRadiobuttonNo.setText("Нет")
        self.InnovationButGroup.addButton(InovationRadiobuttonNo)

        self.dateNoticeReplacement = QtWidgets.QLabel('Дата (период) размещения извещения о закупке', self)
        self.dateNoticeReplacement.setGeometry(QtCore.QRect(70, 610, 550, 50))
        self.dateNoticeReplacement.setFont(font)
        self.butNoticeReplacement = QtWidgets.QPushButton(self)
        self.butNoticeReplacement.setGeometry(QtCore.QRect(580, 620, 38, 30))
        self.butNoticeReplacement.setIcon(calendarButIcon)
        self.butNoticeReplacement.setIconSize(QtCore.QSize(100, 30))
        self.butNoticeReplacement.clicked.connect(
            lambda: self.calen(self.lineNoticeReplacement, self.lineContractExecution, 'start'))
        self.lineNoticeReplacement = QtWidgets.QLineEdit(self)
        self.lineNoticeReplacement.setText("**.**.****")
        self.lineNoticeReplacement.setGeometry(QtCore.QRect(640, 620, 120, 30))
        self.lineNoticeReplacement.setReadOnly(True)
        self.lineNoticeReplacement.setFont(font)
        self.lineNoticeReplacement.setToolTip('Нажмите на иконку календаря и\nвыберите нужную дату')

        self.dateContractExecution = QtWidgets.QLabel('Срок исполнения договора', self)
        self.dateContractExecution.setGeometry(QtCore.QRect(70, 660, 550, 50))
        self.dateContractExecution.setFont(font)
        self.butContractExecution = QtWidgets.QPushButton(self)
        self.butContractExecution.setGeometry(QtCore.QRect(580, 673, 38, 30))
        self.butContractExecution.setIcon(calendarButIcon)
        self.butContractExecution.setIconSize(QtCore.QSize(100, 30))
        self.butContractExecution.clicked.connect(
            lambda: self.calen(self.lineContractExecution, self.lineNoticeReplacement, 'stop'))
        self.lineContractExecution = QtWidgets.QLineEdit(self)
        self.lineContractExecution.setText("**.**.****")
        self.lineContractExecution.setGeometry(QtCore.QRect(640, 673, 120, 30))
        self.lineContractExecution.setReadOnly(True)
        self.lineContractExecution.setFont(font)
        self.lineContractExecution.setToolTip('Нажмите на иконку календаря и\nвыберите нужную дату')

        self.ElPurchase = QtWidgets.QLabel('Закупка в электронной форме', self)
        self.ElPurchase.setGeometry(QtCore.QRect(70, 710, 500, 50))
        self.ElPurchase.setFont(font)
        self.ElPurchaseButGroup = QtWidgets.QButtonGroup(self)
        ElPurchaseRadiobuttonYes = QtWidgets.QRadioButton(self)
        ElPurchaseRadiobuttonYes.setGeometry((QtCore.QRect(590, 710, 100, 50)))
        ElPurchaseRadiobuttonYes.setText("Да")
        self.ElPurchaseButGroup.addButton(ElPurchaseRadiobuttonYes)
        ElPurchaseRadiobuttonNo = QtWidgets.QRadioButton(self)
        ElPurchaseRadiobuttonNo.setGeometry((QtCore.QRect(730, 710, 100, 50)))
        ElPurchaseRadiobuttonNo.setText("Нет")
        self.ElPurchaseButGroup.addButton(ElPurchaseRadiobuttonNo)

        self.TypeZ = QtWidgets.QLabel('Тип закупки', self)
        self.TypeZ.setGeometry(QtCore.QRect(70, 760, 500, 50))
        self.TypeZ.setFont(font)
        self.lineTypeZ = QtWidgets.QLineEdit('Планируемая / Фактическая закупка', self)
        self.lineTypeZ.setGeometry(QtCore.QRect(580, 765, 400, 40))
        self.lineTypeZ.setFont(font)
        self.lineTypeZ.setClearButtonEnabled(1)

        self.Customer = QtWidgets.QLabel('Заказчик', self)
        self.Customer.setGeometry(QtCore.QRect(70, 815, 200, 50))
        self.Customer.setFont(font)
        font.setPointSize(9)
        self.lineCustomer = QtWidgets.QLineEdit(
            'ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ ВЫСШЕГО ОБРАЗОВАНИЯ "НАЦИОНАЛЬНЫЙ ИССЛЕДОВАТЕЛЬСКИЙ ЯДЕРНЫЙ УНИВЕРСИТЕТ "МИФИ"',
            self)
        self.lineCustomer.setGeometry(QtCore.QRect(230, 820, 750, 40))
        self.lineCustomer.setFont(font)
        self.lineCustomer.setClearButtonEnabled(1)
        font.setPointSize(11)

        self.buttom = QtWidgets.QPushButton(self)
        self.buttom.setGeometry((QtCore.QRect(250, 900, 100, 50)))
        self.buttom.setText('OK')

        self.buttom.clicked.connect(lambda: self.fill_and_insert())

        self.BackButtom = QtWidgets.QPushButton('Back', self)
        self.BackButtom.setGeometry(600, 900, 100, 50)
        self.BackButtom.clicked.connect(lambda: self.destroy())

    def fill_dict_data(self):

        def button_value(button):
            if button.buttons()[0].isChecked():
                return "1"
            elif button.buttons()[1].isChecked():
                return "0"
            return ""

        self.dict_data = {}  # Объявляем словарь
        self.dict_data["pn_ext_ident"] = [self.line_pn_ext_ident.text()]
        self.dict_data["ContractCurrency"] = [self.lineContractCurrency.currentText()]
        self.dict_data["PaymentAmount"] = [self.linePaymentAmount.text()]
        self.dict_data["TransCurrency"] = [self.lineTransCurrency.currentText()]
        self.dict_data["AttractionValue"] = [self.lineAttractionValue.text()]
        self.dict_data["Value"] = [self.lineValue.text()]
        self.dict_data["MeasureUnit"] = [self.lineMeasureUnit.text()]
        self.dict_data["SmallBusinessButGroup"] = [button_value(self.SmallBusinessButGroup)]
        self.dict_data["ExcludeButGroup"] = [button_value(self.ExcludeButGroup)]
        self.dict_data["InnovationButGroup"] = [button_value(self.InnovationButGroup)]
        self.dict_data["NoticeReplacement"] = [self.lineNoticeReplacement.text()]
        self.dict_data["ContractExecution"] = [self.lineContractExecution.text()]
        self.dict_data["ElPurchaseButGroup"] = [button_value(self.ElPurchaseButGroup)]
        self.dict_data["TypeZ"] = [self.lineTypeZ.text()]
        self.dict_data["Customer"] = [self.lineCustomer.text()]

        if not self.check_exist():
            return False

        def date_formatting(date):
            import datetime
            date += " 12:01:10"
            date = datetime.datetime.strptime(date, "%d.%m.%Y %H:%M:%S")
            date = date.strftime("%Y-%m-%d %H:%M:%S")
            return date

        conn = psycopg2.connect(dbname='data', user='postgres',
                                password='12345', host='localhost')  # Подключение к бд
        cursor = conn.cursor()  # Специальный объект, который делает запросы и получает их результаты

        pn_ext_ident = self.dict_data["pn_ext_ident"][0]

        query = "SELECT id_plan_position from nm_plan_positions where n_ext_ident = %s" % pn_ext_ident
        cursor.execute(query)
        count_ID_Plan_Position = len(cursor.fetchall())

        if count_ID_Plan_Position != 0:
            query = "SELECT id_contract_currency from nm_plan_positions where n_ext_ident = %s" % pn_ext_ident
            cursor.execute(query)
            value_ID_ContractCurrency = cursor.fetchall()
            if value_ID_ContractCurrency[-1][0] is not None:
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Форма 2 для данной позиции плана уже заполнена")
                mess.exec()
                return False
        else:
            mess = QtWidgets.QMessageBox(self)
            mess.setText("Позиция плана с данным Внешним идентификатором позиции еще не создана")
            mess.exec()
            return False

        query = "SELECT id_plan_position from nm_plan_positions where n_ext_ident = %s" % \
                self.dict_data["pn_ext_ident"][0]
        cursor.execute(query)
        ID_Plan_Position = cursor.fetchall()[0]

        query = "SELECT id_currency from nm_currencies where v_name = '%s'" % self.dict_data["ContractCurrency"][0]
        cursor.execute(query)
        ID_ContractCurrency = cursor.fetchall()[0]

        PaymentAmount = self.dict_data["PaymentAmount"][0]

        query = "SELECT id_currency from nm_currencies where v_name = '%s'" % self.dict_data["TransCurrency"][0]
        cursor.execute(query)
        ID_TransCurrency = cursor.fetchall()[0]

        AttractionValue = self.dict_data["AttractionValue"][0]
        Value = self.dict_data["Value"][0]

        query = "SELECT id_measure_unit from nm_measure_units where v_name = '%s'" % self.dict_data["MeasureUnit"][
            0]
        cursor.execute(query)
        ID_MeasureUnit = cursor.fetchall()[0]

        SmallBusinessButGroup = self.dict_data["SmallBusinessButGroup"][0]

        ExcludeButGroup = self.dict_data["ExcludeButGroup"][0]

        InnovationButGroup = self.dict_data["InnovationButGroup"][0]

        NoticeReplacement = date_formatting(self.dict_data["NoticeReplacement"][0])
        ContractExecution = date_formatting(self.dict_data["ContractExecution"][0])

        ElPurchaseButGroup = self.dict_data["ElPurchaseButGroup"][0]

        TypeZ = self.dict_data["TypeZ"][0]
        Customer = self.dict_data["Customer"][0]

        self.dict_data_f2_db = {}
        self.dict_data_f2_db["pn_ext_ident"] = [pn_ext_ident]
        self.dict_data_f2_db["ID_Plan_Position"] = [ID_Plan_Position]
        self.dict_data_f2_db["ID_Contract_Currency"] = [ID_ContractCurrency]
        self.dict_data_f2_db["PaymentAmount"] = [PaymentAmount]
        self.dict_data_f2_db["ID_TransCurrency"] = [ID_TransCurrency]
        self.dict_data_f2_db["AttractionValue"] = [AttractionValue]
        self.dict_data_f2_db["Value"] = [Value]
        self.dict_data_f2_db["ID_MeasureUnit"] = [ID_MeasureUnit]
        self.dict_data_f2_db["SmallBusinessButGroup"] = [SmallBusinessButGroup]
        self.dict_data_f2_db["ExcludeButGroup"] = [ExcludeButGroup]
        self.dict_data_f2_db["InnovationButGroup"] = [InnovationButGroup]
        self.dict_data_f2_db["NoticeReplacement"] = [NoticeReplacement]
        self.dict_data_f2_db["ContractExecution"] = [ContractExecution]
        self.dict_data_f2_db["ElPurchaseButGroup"] = [ElPurchaseButGroup]
        self.dict_data_f2_db["TypeZ"] = [TypeZ]
        self.dict_data_f2_db["Customer"] = [Customer]

        conn.close()
        return True

    def check_exist(self):
        for key in self.dict_data.keys():
            if len(self.dict_data[key][0]) == 0:
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля!!!")
                mess.exec()
                return False
            elif self.dict_data[key][0] == "**.**.****":
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля!!!")
                mess.exec()
                return False
        return True

    def insert_in_db(self):
        conn = psycopg2.connect(dbname='data', user='postgres',
                                password='12345', host='localhost')
        cursor = conn.cursor()

        query = "call public.save_plan_position(\n" \
                "\tpid_plan_position => {},\n" \
                "\tpid_plan => 1,\n" \
                "\tpid_contract_currency => {},\n" \
                "\tpf_payment_amount => {},\n" \
                "\tpid_trans_currency => {},\n" \
                "\tpf_attraction_value => {},\n" \
                "\tpf_value => {},\n" \
                "\tpid_measure_unit => {},\n" \
                "\tpb_small_business => {},\n" \
                "\tpb_exclude => {},\n" \
                "\tpb_innovation => {},\n" \
                "\tpDT_NOTICE_REPLACEMENT => '{}',\n" \
                "\tpDT_CONTRACT_EXECUTION => '{}',\n" \
                "\tpb_el_purchase => {},\n" \
                "\tpv_customer => '{}',\n" \
                "\tpid_user => 1,\n" \
                "\tpid_region => 1,\n" \
                "\tpn_ext_ident => {},\n" \
                "\tpaction => 2);".format(self.dict_data_f2_db["ID_Plan_Position"][0][0],
                                          self.dict_data_f2_db["ID_Contract_Currency"][0][0],
                                          self.dict_data_f2_db["PaymentAmount"][0],
                                          self.dict_data_f2_db["ID_TransCurrency"][0][0],
                                          self.dict_data_f2_db["AttractionValue"][0],
                                          self.dict_data_f2_db["Value"][0],
                                          self.dict_data_f2_db["ID_MeasureUnit"][0][0],
                                          self.dict_data_f2_db["SmallBusinessButGroup"][0],
                                          self.dict_data_f2_db["ExcludeButGroup"][0],
                                          self.dict_data_f2_db["InnovationButGroup"][0],
                                          self.dict_data_f2_db["NoticeReplacement"][0],
                                          self.dict_data_f2_db["ContractExecution"][0],
                                          self.dict_data_f2_db["ElPurchaseButGroup"][0],
                                          self.dict_data_f2_db["Customer"][0],
                                          self.dict_data_f2_db["pn_ext_ident"][0])

        cursor.execute(query)
        conn.commit()
        conn.close()
        return None

    def fill_and_insert(self):
        if self.fill_dict_data():
            text = "Заполнена Форма 2 для позиции плана с Внешним идентификатором записи = {}.".format(
                self.dict_data_f2_db["pn_ext_ident"][0])
            self.insert_in_db()
            mess = QtWidgets.QMessageBox(self)
            mess.setText(text)
            mess.exec()
            return True
        else:
            return False


class CreatePlanPosition(QtWidgets.QWidget):
    def calen(self, label_1, label_2, start_stop):
        cal = CalendarBox(self, label_1, label_2, start_stop)
        cal.resize(474, 373)
        cal.show()

    def next(self):
        self.nex = CreateF2()
        self.nex.resize(1000, 1000)
        self.nex.show()

    def __init__(self):
        super().__init__()  # возвращает родительский объект CreatePlan с классом
        self.resize(1000, 1000)  # минимальные размеры окна

        global db
        db.setHostName("localhost")
        db.setPort(5432)
        db.setDatabaseName("data")
        db.setUserName('browser')
        db.setPassword("browser")
        db.open()
        font = QtGui.QFont()  # метод для создания GUI
        font.setPointSize(11)

        calendarPixMap = QtGui.QPixmap("GovBuy/design/icons/img_52517.png")
        calendarButIcon = QtGui.QIcon(calendarPixMap)

        self.pn_ext_ident = QtWidgets.QLabel('Внешний идентификатор позиции', self)
        self.pn_ext_ident.setGeometry(QtCore.QRect(120, 10, 300, 40))
        self.pn_ext_ident.setFont(font)
        self.line_pn_ext_ident = QtWidgets.QLineEdit(self)
        self.line_pn_ext_ident.setGeometry(QtCore.QRect(460, 10, 500, 40))
        self.line_pn_ext_ident.setFont(font)
        self.line_pn_ext_ident.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("[1-9]\d+(\.|,)\d+")))
        self.line_pn_ext_ident.setClearButtonEnabled(1)
        self.line_pn_ext_ident.setToolTip('Начинайте вводить внешний идентификатор позиции')

        self.OKVED2 = QtWidgets.QLabel('Код по ОКВЭД2', self)
        self.OKVED2.setGeometry(QtCore.QRect(120, 70, 300, 40))
        self.OKVED2.setFont(font)
        self.lineOKVED2 = LineEditCompleter(self, db, "SELECT v_okved2||' '||v_name from nm_okved2")
        self.lineOKVED2.setGeometry(QtCore.QRect(460, 70, 500, 40))
        self.lineOKVED2.setFont(font)
        self.lineOKVED2.setToolTip('Начинайте вводить код ОКВЭД2')

        self.OKPD2 = QtWidgets.QLabel('ОКПД2', self)  # 4 digits
        self.OKPD2.setGeometry(QtCore.QRect(120, 130, 300, 40))
        self.OKPD2.setFont(font)
        self.lineOKPD2 = LineEditCompleter(self, db, "SELECT v_okpd2||' '||v_name from nm_okpd2")
        self.lineOKPD2.setGeometry(QtCore.QRect(460, 130, 500, 40))
        self.lineOKPD2.setFont(font)
        self.lineOKPD2.setToolTip('Начинайте вводить код ОКПД2')

        self.object = QtWidgets.QLabel('Предмет закупки', self)
        self.object.setGeometry(QtCore.QRect(120, 190, 300, 40))
        self.object.setFont(font)
        self.lineObject = QtWidgets.QLineEdit(self)
        self.lineObject.setGeometry(QtCore.QRect(460, 190, 500, 40))
        self.lineObject.setFont(font)
        self.lineObject.setClearButtonEnabled(1)

        self.deliveryPlace = QtWidgets.QLabel('Место поставки товара', self)
        self.deliveryPlace.setGeometry(QtCore.QRect(120, 250, 300, 40))
        self.deliveryPlace.setFont(font)
        self.lineDeliveryPlace = ComboBoxCompleter(self, db, "select v_name from nm_regions;")
        self.lineDeliveryPlace.setGeometry(QtCore.QRect(460, 250, 500, 40))
        self.lineDeliveryPlace.setFont(font)
        self.lineDeliveryPlace.setToolTip('Выберите субъект Федерации из списка')

        self.startPrice = QtWidgets.QLabel('Начальная стоимость', self)
        self.startPrice.setGeometry(QtCore.QRect(120, 310, 300, 40))
        self.startPrice.setFont(font)
        self.lineStartPrice = QtWidgets.QLineEdit(self)
        self.lineStartPrice.setGeometry(QtCore.QRect(460, 310, 500, 40))
        self.lineStartPrice.setFont(font)
        self.lineStartPrice.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("[1-9]\d+(\.|,)\d+")))
        self.lineStartPrice.setClearButtonEnabled(1)
        self.lineStartPrice.setToolTip('Начальная (максимальная) цена договора - (Для долгосрочной \n'
                                       'закупки - общая величина цены договора), руб.')

        self.dateStart = QtWidgets.QLabel('Начало периода размещения\nизвещения о закупке', self)
        self.dateStart.setGeometry(QtCore.QRect(120, 370, 300, 60))
        self.dateStart.setFont(font)
        self.butDataStart = QtWidgets.QPushButton(self)
        self.butDataStart.setGeometry(QtCore.QRect(582, 370, 38, 30))
        self.butDataStart.setIcon(calendarButIcon)
        self.butDataStart.setIconSize(QtCore.QSize(100, 30))
        self.butDataStart.clicked.connect(lambda: self.calen(self.lineDataStart, self.lineDataStop, 'start'))
        self.lineDataStart = QtWidgets.QLineEdit(self)
        self.lineDataStart.setText("**.**.****")
        self.lineDataStart.setGeometry(QtCore.QRect(460, 370, 120, 30))
        self.lineDataStart.setReadOnly(True)
        self.lineDataStart.setFont(font)
        self.lineDataStart.setToolTip('Нажмите на иконку календаря и\nвыберите нужную дату')

        self.dateStop = QtWidgets.QLabel('Дата окончания исполнения\nзакупки', self)
        self.dateStop.setGeometry(QtCore.QRect(120, 440, 300, 60))
        self.dateStop.setFont(font)
        self.butDataStop = QtWidgets.QPushButton(self)
        self.butDataStop.setGeometry(QtCore.QRect(582, 440, 38, 30))
        self.butDataStop.setIcon(calendarButIcon)
        self.butDataStop.setIconSize(QtCore.QSize(100, 30))
        self.butDataStop.clicked.connect(lambda: self.calen(self.lineDataStop, self.lineDataStart, 'stop'))
        self.lineDataStop = QtWidgets.QLineEdit(self)
        self.lineDataStop.setText("**.**.****")
        self.lineDataStop.setGeometry(QtCore.QRect(460, 440, 120, 30))
        self.lineDataStop.setReadOnly(True)
        self.lineDataStop.setFont(font)
        self.lineDataStop.setToolTip('Нажмите на иконку календаря и\nвыберите нужную дату')

        self.purchaseWay = QtWidgets.QLabel('Способ закупки', self)
        self.purchaseWay.setGeometry(QtCore.QRect(120, 510, 300, 40))
        self.purchaseWay.setFont(font)
        self.linePurchaseWay = ComboBoxCompleter(self, db,
                                                 "select v_name from nm_purchase_ways where b_deleted=0")
        self.linePurchaseWay.setGeometry(QtCore.QRect(460, 510, 500, 40))
        self.linePurchaseWay.setFont(font)
        self.linePurchaseWay.setToolTip('Выберите способ из списка')

        self.purchaseCondition = QtWidgets.QLabel('Условия закупки у СМП', self)
        self.purchaseCondition.setGeometry(QtCore.QRect(120, 570, 300, 40))
        self.purchaseCondition.setFont(font)
        self.linePurchaseCondition = ComboBoxCompleter(parent=self, db=db,
                                                       query="select v_name from nm_purchase_conditions")
        self.linePurchaseCondition.setGeometry(QtCore.QRect(460, 570, 500, 40))
        self.linePurchaseCondition.setFont(font)
        self.linePurchaseCondition.setToolTip('Выберите условие закупки из списка')

        self.breakdown = QtWidgets.QLabel('Разбивка оплаты по годам', self)
        self.breakdown.setGeometry(QtCore.QRect(120, 630, 300, 40))
        self.breakdown.setFont(font)
        self.lineBreakdown = QtWidgets.QLineEdit(self)
        self.lineBreakdown.setGeometry(QtCore.QRect(460, 630, 500, 40))
        self.lineBreakdown.setFont(font)
        self.lineBreakdown.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("(\d{4}-[1-9]\d+(((\.|,)\d)|)+\s)+")))
        self.lineBreakdown.setClearButtonEnabled(1)
        self.lineBreakdown.setToolTip('Введите год и цену через тире, а сами записи через пробел\n'
                                      'Пример: год1-цена1 год2-цена2 ...')

        self.financingSource = QtWidgets.QLabel('Источник финансирования', self)
        self.financingSource.setGeometry(QtCore.QRect(120, 690, 300, 40))
        self.financingSource.setFont(font)
        self.lineFinancialSources = ComboBoxCompleter(self, db,
                                                      "select V_name||' '||v_description, id_financing_source from NM_FINANCING_SOURCES;")
        self.lineFinancialSources.setGeometry(460, 690, 500, 40)
        self.lineFinancialSources.setFont(font)
        self.lineFinancialSources.setToolTip('Выберите источник из списка')
        db.close()

        self.contract = QtWidgets.QLabel('Пункты положения о закупках\nМИФИ', self)
        self.contract.setGeometry(QtCore.QRect(120, 740, 310, 60))
        self.contract.setFont(font)
        self.lineContract = QtWidgets.QLineEdit(self)
        self.lineContract.setGeometry(QtCore.QRect(460, 750, 500, 40))
        self.lineContract.setFont(font)
        self.lineContract.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("[а-яА-Я0-9.,-]+")))
        self.lineContract.setClearButtonEnabled(1)
        self.lineContract.setToolTip('Введите пунк положения')
        self.lineContract.text()

        buttom = QtWidgets.QPushButton(self)
        buttom.setGeometry((QtCore.QRect(250, 850, 100, 50)))
        buttom.setText('OK')

        buttom.clicked.connect(lambda: self.fill_and_insert())

        self.BackButtom = QtWidgets.QPushButton('Back', self)
        self.BackButtom.setGeometry(600, 850, 100, 50)

        self.NextButtom = QtWidgets.QPushButton('Next', self)
        self.NextButtom.setGeometry(850, 800, 100, 50)

        self.NextButtom.clicked.connect(lambda: self.next())

    def fill_dict_data(self):

        self.dict_data = {}  # Объявляем словарь
        # Для поля ОКВЭД2 создаем соответствующий ключ в словаре,
        # а введенное в поле значение сохраняем в список
        # для остальных полей аналогично

        self.dict_data["pn_ext_ident"] = [self.line_pn_ext_ident.text()]
        self.dict_data["OKVED2"] = [self.lineOKVED2.text()]
        self.dict_data["OKPD2"] = [self.lineOKPD2.text()]
        self.dict_data["Object"] = [self.lineObject.text()]
        self.dict_data["DeliveryPlace"] = [self.lineDeliveryPlace.currentText()]
        self.dict_data["StartPrice"] = [self.lineStartPrice.text()]
        self.dict_data["DataStart"] = [self.lineDataStart.text()]
        self.dict_data["DataStop"] = [self.lineDataStop.text()]
        self.dict_data["PurchaseWay"] = [self.linePurchaseWay.currentText()]
        self.dict_data["PurchaseCondition"] = [self.linePurchaseCondition.currentText()]
        self.dict_data["Breakdown"] = [self.lineBreakdown.text()]
        self.dict_data["FinancialSources"] = [self.lineFinancialSources.currentText()]
        self.dict_data["Contract"] = [self.lineContract.text()]

        if not self.check_exist():
            return False

        def date_formatting(date):
            import datetime
            date += " 12:01:10"
            date = datetime.datetime.strptime(date, "%d.%m.%Y %H:%M:%S")
            date = date.strftime("%Y-%m-%d %H:%M:%S")
            return date

        conn = psycopg2.connect(dbname='data', user='postgres',
                                password='12345', host='localhost')  # Подключение к бд
        cursor = conn.cursor()  # Специальный объект, который делает запросы и получает их результаты

        pn_ext_ident = self.dict_data["pn_ext_ident"][0]

        query = "SELECT id_plan_position from nm_plan_positions where n_ext_ident = %s" % pn_ext_ident
        cursor.execute(query)
        count_ID_Plan_Position = len(cursor.fetchall())

        if count_ID_Plan_Position != 0:
            mess = QtWidgets.QMessageBox(self)
            mess.setText("Позиция плана с данным Внешним идентификатором позиции уже создана")
            mess.exec()
            return False

        query = "SELECT id_okved2 from nm_okved2 where v_okved2 = '%s'" % self.dict_data["OKVED2"][0].split(' ')[0]
        cursor.execute(query)
        ID_OKVED = cursor.fetchall()[0]

        query = "SELECT id_okpd2 from nm_okpd2 where v_okpd2 = '%s'" % self.dict_data["OKPD2"][0].split(' ')[0]
        cursor.execute(query)
        ID_OKPD = cursor.fetchall()[0]

        Object = self.dict_data["Object"][0]
        DeliveryPlace = self.dict_data["DeliveryPlace"][0]
        StartPrice = self.dict_data["StartPrice"][0]
        DataStart = date_formatting(self.dict_data["DataStart"][0])
        DataStop = date_formatting(self.dict_data["DataStop"][0])

        query = "SELECT id_purchase_way from nm_purchase_ways where v_name = '%s'" % self.dict_data["PurchaseWay"][0]
        cursor.execute(query)
        ID_PurchaseWay = cursor.fetchall()[0]

        query = "SELECT id_purchase_condition from nm_purchase_conditions where v_name = '%s'" % \
                self.dict_data["PurchaseCondition"][0]
        cursor.execute(query)
        ID_PurchaseCondition = cursor.fetchall()[0]

        # Разбивку оплаты по годам (Breakdown или payment_breakdown) пока не делаю, потому что еще с этим не разобрались

        query = "SELECT id_financing_source from NM_FINANCING_SOURCES where V_name = '%s'" % \
                self.dict_data["FinancialSources"][0].split(" ")[0]
        cursor.execute(query)
        ID_FinancialSources = cursor.fetchall()[0]

        # Пункты положения о закупках (Contract) пока не делаю, потому что еще с этим не разобрались

        self.dict_data_f1_db = {}
        self.dict_data_f1_db["pn_ext_ident"] = [pn_ext_ident]
        self.dict_data_f1_db["ID_OKVED2"] = [ID_OKVED]
        self.dict_data_f1_db["ID_OKPD2"] = [ID_OKPD]
        self.dict_data_f1_db["Object"] = [Object]
        self.dict_data_f1_db["DeliveryPlace"] = [DeliveryPlace]
        self.dict_data_f1_db["StartPrice"] = [StartPrice]
        self.dict_data_f1_db["DataStart"] = [DataStart]
        self.dict_data_f1_db["DataStop"] = [DataStop]
        self.dict_data_f1_db["ID_PurchaseWay"] = [ID_PurchaseWay]
        self.dict_data_f1_db["ID_PurchaseCondition"] = [ID_PurchaseCondition]
        # self.dict_data_f1_db["ID_Breakdown"] пока нету
        self.dict_data_f1_db["ID_FinancialSources"] = [ID_FinancialSources]
        # self.dict_data_f1_db["Contract"] пока нету

        conn.close()
        return True

    def check_exist(self):
        for key in self.dict_data.keys():
            if len(self.dict_data[key][0]) == 0:
                # print(self.nex.lineContractCurrency.currentText())
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля!!!")
                mess.exec()
                return False
            elif self.dict_data[key][0] == "**.**.****":
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля!!!")
                mess.exec()
                return False
        return True

    def insert_in_db(self):
        conn = psycopg2.connect(dbname='data', user='postgres',
                                password='12345', host='localhost')
        cursor = conn.cursor()

        query = "call public.save_plan_position(\n" \
                "\tpid_plan => 1,\n" \
                "\tpv_position_status => 'A',\n" \
                "\tpid_okved2 => {},\n" \
                "\tpid_okpd2 => {},\n" \
                "\tpv_purchase_object => '{}',\n" \
                "\tpv_delivery_place => '{}',\n" \
                "\tpf_start_max_price => {},\n" \
                "\tpdt_purchase_start => '{}',\n" \
                "\tpdt_purchase_stop => '{}',\n" \
                "\tpid_purchase_way => {},\n" \
                "\tpid_purchase_condition => {},\n" \
                "\tpid_payment_breakdown => null,\n" \
                "\tpid_financing_source => {},\n" \
                "\tpid_user => 1,\n" \
                "\tpid_region => 1,\n" \
                "\tpn_ext_ident => {},\n" \
                "\tpaction => 1);".format(self.dict_data_f1_db["ID_OKVED2"][0][0],
                                          self.dict_data_f1_db["ID_OKPD2"][0][0],
                                          self.dict_data_f1_db["Object"][0], self.dict_data_f1_db["DeliveryPlace"][0],
                                          self.dict_data_f1_db["StartPrice"][0], self.dict_data_f1_db["DataStart"][0],
                                          self.dict_data_f1_db["DataStop"][0],
                                          self.dict_data_f1_db["ID_PurchaseWay"][0][0],
                                          self.dict_data_f1_db["ID_PurchaseCondition"][0][0],
                                          self.dict_data_f1_db["ID_FinancialSources"][0][0],
                                          self.dict_data_f1_db["pn_ext_ident"][0])

        cursor.execute(query)
        conn.commit()
        conn.close()
        return None

    def fill_and_insert(self):
        if self.fill_dict_data():
            text = "Создана позиция плана с Внешним идентификатором записи = {}. Форма 1 заполнена.".format(
                self.dict_data_f1_db["pn_ext_ident"][0])
            self.insert_in_db()
            mess = QtWidgets.QMessageBox(self)
            mess.setText(text)
            mess.exec()
            return True
        else:
            return False


class ChangePlan(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        p = QtWidgets.QLabel('Change Plan', self)
        p.setGeometry(QtCore.QRect(160, 150, 381, 81))
        self.BackButtom = QtWidgets.QPushButton('Back', self)
        self.BackButtom.setGeometry(300, 400, 300, 50)


class UploadPlan(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        #p = QtWidgets.QLabel('Change Plan', self)
        #p.setGeometry(QtCore.QRect(160, 150, 381, 81))
        self.BackButtom = QtWidgets.QPushButton('Back', self)
        self.BackButtom.setGeometry(300, 400, 300, 50)

        self.buttom1 = QtWidgets.QPushButton(self)
        self.buttom1.setGeometry((QtCore.QRect(100, 200, 250, 75)))
        self.buttom1.setText('Загрузить')
        self.buttom1.setFont(QtGui.QFont("Times", 14, QtGui.QFont.Bold))

        self.buttom2 = QtWidgets.QPushButton(self)
        self.buttom2.setGeometry((QtCore.QRect(400, 200, 250, 75)))
        self.buttom2.setText('Обновить')
        self.buttom2.setFont(QtGui.QFont("Times", 14, QtGui.QFont.Bold))

        self.buttom3 = QtWidgets.QPushButton(self)
        self.buttom3.setGeometry((QtCore.QRect(700, 200, 250, 75)))
        self.buttom3.setText('Удалить')
        self.buttom3.setFont(QtGui.QFont("Times", 14, QtGui.QFont.Bold))

        self.value = str(self.buttom1.text())

        self.buttom1.clicked.connect(lambda: self.get_Task('update'))
        self.buttom1.clicked.connect(self.openFileNameDialog)

        self.buttom2.clicked.connect(lambda: self.get_Task('upload'))
        self.buttom2.clicked.connect(self.openFileNameDialog)

        self.buttom3.clicked.connect(lambda: self.get_Task('delete'))
        self.buttom3.clicked.connect(self.openFileNameDialog)

    def get_Task(self, value):
        self.value = value
        # print(self.value)

    def openFileNameDialog(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        fileName, _ = QtWidgets.QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                            "All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            res = self.read_xls(fileName)
            self.dialog = Second(res, self.value)
            self.dialog.setAttribute(QtCore.Qt.WA_DeleteOnClose)
            self.dialog.show()

    def read_xls(self, fileName):
        wb = load_workbook(fileName)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        table = np.zeros(13).reshape(1, 13)
        array = np.zeros(13).reshape(1, 13)
        i = 19
        cycle = True
        count = 0
        while cycle == True:
            check_array = [[str(type(cell.value)) for cell in col] for col in sheet['B' + str(i):'N' + str(i)]]
            true_array = [[cell.value for cell in col] for col in sheet['B' + str(i):'N' + str(i)]]
            unique, counts = np.unique(check_array, return_counts=True)
            print(unique, counts)
            for j in range(unique[np.where(counts == max(counts))].shape[0]):
                if unique[np.where(counts == max(counts))][j] != "<class 'NoneType'>" and counts[
                    np.where(counts == max(counts))[0][j]] >= 6:
                    array = np.append(array, true_array, axis=0)
                elif counts[np.where(counts == max(counts))[0][j]] == 13:
                    count += 1
                    if count > 5:
                        cycle = False
            i += 1
        if array[1][0] == '1':
            return(array[2:])
        else:
            return(array[1:])


class Second(QtWidgets.QDialog):
    def calen(self, label_1, label_2, start_stop):
        cal = CalendarBox(self, label_1, label_2, start_stop)
        cal.resize(474, 373)
        cal.show()

    def check_exist(self):
        pass

    def __init__(self, text, value, parent=None):
        super(Second, self).__init__(parent=parent)
        self.resize(500, 400)

        global db
        db.setHostName("localhost")
        db.setPort(5432)
        db.setDatabaseName("postgres")
        db.setUserName('browser')
        db.setPassword("browser")
        db.open()

        self.text = text
        self.text_new = self.text
        self.task = value
        self.layout = QtWidgets.QHBoxLayout(self)
        self.scrollArea = QtWidgets.QScrollArea(self)
        self.scrollArea.setWidgetResizable(True)
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.gridLayout = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.layout.addWidget(self.scrollArea)

        self.values = ['Порядковый номер', 'ОКВЭД2                              ', 'ОКПД2                              ', 'Предмет закупки', 'Место поставки товара', 'Начальная Цена',
                       'Начальная дата', '', 'Конечная дата', '',
                       'Способ закупки', 'Условия закупки у СМП', 'Разбивка по годам', 'Источник финансирования',
                       'Положения о закупке']

        for i in range(len(self.values)):
            for j in range(1, self.text.shape[0] * 2, 2):
                self.le = QtWidgets.QLabel(self)
                self.le.setText(self.values[i])
                # self.le.move(130, 22)
                self.gridLayout.addWidget(self.le, j, i)
        self.datestartdict = {}
        self.datebut1dict = {}
        self.dateenddict = {}
        self.datebut2dict = {}
        for j in range(2, self.text.shape[0] * 2 + 1, 2):
            self.le = QtWidgets.QLineEdit(self)
            self.lineOKVED2 = LineEditCompleter(self, db, "SELECT v_okved2||' '||v_name from nm_okved2")

            self.lineOKPD2 = LineEditCompleter(self, db, "SELECT v_okpd2||' '||v_name from nm_okpd2")

            self.lineObject = QtWidgets.QLineEdit(self)

            self.lineDeliveryPlace = ComboBoxCompleter(self, db, "select v_name, id_region from nm_regions;")

            self.lineDeliveryPlace.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
            self.lineDeliveryPlace.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToMinimumContentsLengthWithIcon)

            self.lineStartPrice = QtWidgets.QLineEdit(self)
            self.lineStartPrice.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("[1-9]\d+(\.|,)\d+")))

            #self.butDataStart1 = QtWidgets.QPushButton(self)
            #self.butDataStart1.clicked.connect(lambda: self.calen(self.lineDataStart, self.lineDataStop, 'start'))

            #self.lineDataStart = QtWidgets.QLineEdit(self)
            #self.lineDataStart.setText("**.**.****")
            #self.lineDataStart.setReadOnly(True)

            #self.butDataStart2 = QtWidgets.QPushButton(self)
            #self.butDataStart2.clicked.connect(lambda: self.calen(self.lineDataStop, self.lineDataStart, 'stop'))

            #self.lineDataStop = QtWidgets.QLineEdit(self)
            #self.lineDataStop.setText("**.**.****")
            #self.lineDataStop.setReadOnly(True)

            self.datebut1dict[str(j)] = QtWidgets.QPushButton(self)
            self.datebut1dict[str(j)].clicked.connect(lambda: self.calen(self.datestartdict[str(j)], self.dateenddict[str(j)], 'start'))

            self.datestartdict[str(j)] = QtWidgets.QLineEdit(self)
            self.datestartdict[str(j)].setText("**.**.****")
            #self.datestartdict[str(j)].setReadOnly(True)


            self.datebut2dict[str(j)] = QtWidgets.QPushButton(self)
            self.datebut2dict[str(j)].clicked.connect(lambda: self.calen(self.dateenddict[str(j)], self.datestartdict[str(j)], 'stop'))

            self.dateenddict[str(j)] = QtWidgets.QLineEdit(self)
            self.dateenddict[str(j)].setText("**.**.****")
            #self.dateenddict[str(j)].setReadOnly(True)





            self.linePurchaseWay = ComboBoxCompleter(self, db,
                                                     "select v_name, id_purchase_way from nm_purchase_ways where b_deleted=0")

            self.linePurchaseWay.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
            self.linePurchaseWay.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToMinimumContentsLengthWithIcon)

            self.linePurchaseCondition = ComboBoxCompleter(parent=self, db=db,
                                                           query="select v_name, id_purchase_condition from nm_purchase_conditions")

            self.linePurchaseCondition.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
            self.linePurchaseCondition.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToMinimumContentsLengthWithIcon)

            self.lineBreakdown = QtWidgets.QLineEdit(self)

            self.lineBreakdown.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("(\d{4}-[1-9]\d+(((\.|,)\d)|)+\s)+")))

            self.lineFinancialSources = ComboBoxCompleter(self, db,
                                                          "select V_name||' '||v_description, id_financing_source from NM_FINANCING_SOURCES;")

            self.lineFinancialSources.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
            self.lineFinancialSources.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToMinimumContentsLengthWithIcon)

            self.lineContract = QtWidgets.QLineEdit(self)

            self.lineContract.setValidator(QtGui.QRegExpValidator(QtCore.QRegExp("[а-яА-Я0-9.,-]+")))

            index_a = int((j) / 2)
            self.le.setText(str(self.text[index_a - 1, 0]))
            self.lineOKVED2.setText(str(self.text[index_a - 1, 1]))
            self.lineOKPD2.setText(str(self.text[index_a - 1, 2]))
            self.lineObject.setText(str(self.text[index_a - 1, 3]))

            index = self.lineDeliveryPlace.findText(str(self.text[index_a - 1, 4]), QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.lineDeliveryPlace.setCurrentIndex(index)

            self.lineStartPrice.setText(str(self.text[index_a - 1, 5]))

            self.datestartdict[str(j)].setText(str(self.text[index_a - 1, 6]).strip())
            self.dateenddict[str(j)].setText(str(self.text[index_a - 1, 7]).strip())


            index = self.linePurchaseWay.findText(str(self.text[index_a - 1, 8]), QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.linePurchaseWay.setCurrentIndex(index)

            index = self.linePurchaseCondition.findText(str(self.text[index_a - 1, 9]), QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.linePurchaseCondition.setCurrentIndex(index)

            self.lineBreakdown.setText(str(self.text[index_a - 1, 10]))

            index = self.lineFinancialSources.findText(str(self.text[index_a - 1, 11]), QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.lineFinancialSources.setCurrentIndex(index)

            self.lineContract.setText(str(self.text[index_a - 1, 12]))

            # self.le.move(130, 22)
            self.gridLayout.addWidget(self.le, j, 0)
            self.gridLayout.addWidget(self.lineOKVED2, j, 1)
            self.gridLayout.addWidget(self.lineOKPD2, j, 2)
            self.gridLayout.addWidget(self.lineObject, j, 3)
            self.gridLayout.addWidget(self.lineDeliveryPlace, j, 4)
            self.gridLayout.addWidget(self.lineStartPrice, j, 5)
            #self.gridLayout.addWidget(self.butDataStart1, j, 7)
            #self.gridLayout.addWidget(self.lineDataStart, j, 6)
            #self.gridLayout.addWidget(self.butDataStart2, j, 9)
            #self.gridLayout.addWidget(self.lineDataStop, j, 8)
            self.gridLayout.addWidget(self.datebut1dict[str(j)], j, 7)
            self.gridLayout.addWidget(self.datestartdict[str(j)], j, 6)
            self.gridLayout.addWidget(self.datebut2dict[str(j)], j, 9)
            self.gridLayout.addWidget(self.dateenddict[str(j)], j, 8)
            self.gridLayout.addWidget(self.linePurchaseWay, j, 10)
            self.gridLayout.addWidget(self.linePurchaseCondition, j, 11)
            self.gridLayout.addWidget(self.lineBreakdown, j, 12)
            self.gridLayout.addWidget(self.lineFinancialSources, j, 13)
            self.gridLayout.addWidget(self.lineContract, j, 14)



        self.button = QtWidgets.QPushButton('OK', self)
        self.gridLayout.addWidget(self.button, j + 1, i)
        self.button.clicked.connect(self.collect_data)

        self.BackButtom = QtWidgets.QPushButton('back', self)
        self.gridLayout.addWidget(self.BackButtom, j + 1, 0)
        self.BackButtom.clicked.connect(self.close)
        self.BackButtom.clicked.connect(self.openFileNameDialog)

    def collect_data(self):
        # conn = psycopg2.connect(dbname='postgres', user='postgres',
        #                password='******', host='localhost')
        # cursor = conn.cursor()
        for j in range(2, self.text.shape[0] * 2 + 1, 2):
            index = int((j) / 2)
            # self.text_new[index-1,i]=self.gridLayout.itemAtPosition(j, i).widget().text()
            if self.fill_dict_data(j):
                text = "Заполнена Форма 2 для позиции плана с Внешним идентификатором записи = {}.".format(
                    self.gridLayout.itemAtPosition(self.j, 0).widget().text())
                self.insert_in_db()
                mess = QtWidgets.QMessageBox(self)
                mess.setText(text)
                mess.exec()
            else:
                mess = QtWidgets.QMessageBox(self)
                mess.setText('ошибка в строке %s' %(int(self.j / 2)))
                mess.exec()
            # self.check_exist(j)


        # return(self.text_new, self.task)

    def fill_dict_data(self, j):
        self.j = j
        self.dict_data = {}
        self.dict_data["pn_ext_ident"] = [self.gridLayout.itemAtPosition(self.j, 0).widget().text()]
        self.dict_data["OKVED2"] = [self.gridLayout.itemAtPosition(self.j, 1).widget().text()]
        self.dict_data["OKPD2"] = [self.gridLayout.itemAtPosition(self.j, 2).widget().text()]
        self.dict_data["Object"] = [self.gridLayout.itemAtPosition(self.j, 3).widget().text()]
        self.dict_data["DeliveryPlace"] = [self.gridLayout.itemAtPosition(self.j, 4).widget().currentText()]
        self.dict_data["StartPrice"] = [self.gridLayout.itemAtPosition(self.j, 5).widget().text()]
        self.dict_data["DataStart"] = [self.gridLayout.itemAtPosition(self.j, 6).widget().text()]
        self.dict_data["DataStop"] = [self.gridLayout.itemAtPosition(self.j, 8).widget().text()]
        self.dict_data["PurchaseWay"] = [self.gridLayout.itemAtPosition(self.j, 10).widget().currentText()]
        self.dict_data["PurchaseCondition"] = [self.gridLayout.itemAtPosition(self.j, 11).widget().currentText()]
        self.dict_data["Breakdown"] = [self.gridLayout.itemAtPosition(self.j, 12).widget().text()]
        self.dict_data["FinancialSources"] = [self.gridLayout.itemAtPosition(self.j, 13).widget().currentText()]
        self.dict_data["Contract"] = [self.gridLayout.itemAtPosition(self.j, 14).widget().text()]

        print(self.dict_data["OKVED2"][0].split(' ')[0], self.dict_data["PurchaseWay"][0],
              self.dict_data["FinancialSources"][0].split(" ")[0])

        #if not self.check_exist(self.j):
        #    return False

        for key in self.dict_data.keys():
            if len(self.dict_data[key][0]) == 0:
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля в строчке %s!!!" % (int(self.j / 2)))
                mess.exec()
                return False
            elif self.dict_data[key][0] == "**.**.****":
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля в строчке %s!!!" % (int(self.j / 2)))
                mess.exec()
                return False

        def date_formatting(date):
            import datetime
            date += " 12:01:10"
            date = datetime.datetime.strptime(date, "%d.%m.%Y %H:%M:%S")
            date = date.strftime("%Y-%m-%d %H:%M:%S")
            return date

        conn = psycopg2.connect(dbname='postgres', user='postgres',
                                password='300771Den', host='localhost')  # Подключение к бд
        cursor = conn.cursor()  # Специальный объект, который делает запросы и получает их результаты

        pn_ext_ident = self.dict_data["pn_ext_ident"][0]

        query = "SELECT id_plan_position from nm_plan_positions where n_ext_ident = %s" % pn_ext_ident
        cursor.execute(query)
        count_ID_Plan_Position = len(cursor.fetchall())

        if count_ID_Plan_Position != 0:
            mess = QtWidgets.QMessageBox(self)
            mess.setText("Позиция плана с данным Внешним идентификатором позиции уже создана")
            mess.exec()
            return False

        query = "SELECT id_okved2 from nm_okved2 where v_okved2 = '%s'" % self.dict_data["OKVED2"][0].split(' ')[0]
        cursor.execute(query)
        ID_OKVED = cursor.fetchall()[0]

        query = "SELECT id_okpd2 from nm_okpd2 where v_okpd2 = '%s'" % self.dict_data["OKPD2"][0].split(' ')[0]
        cursor.execute(query)
        ID_OKPD = cursor.fetchall()[0]

        Object = self.dict_data["Object"][0]
        DeliveryPlace = self.dict_data["DeliveryPlace"][0]
        StartPrice = self.dict_data["StartPrice"][0]
        DataStart = date_formatting(self.dict_data["DataStart"][0])
        DataStop = date_formatting(self.dict_data["DataStop"][0])

        query = "SELECT id_purchase_way from nm_purchase_ways where v_name = '%s'" % self.dict_data["PurchaseWay"][0]
        cursor.execute(query)
        ID_PurchaseWay = cursor.fetchall()[0]

        query = "SELECT id_purchase_condition from nm_purchase_conditions where v_name = '%s'" % \
                self.dict_data["PurchaseCondition"][0]
        cursor.execute(query)
        ID_PurchaseCondition = cursor.fetchall()[0]

        # Разбивку оплаты по годам (Breakdown или payment_breakdown) пока не делаю, потому что еще с этим не разобрались

        query = "SELECT id_financing_source from NM_FINANCING_SOURCES where V_name = '%s'" % \
                self.dict_data["FinancialSources"][0].split(" ")[0]
        cursor.execute(query)
        ID_FinancialSources = cursor.fetchall()[0]

        # Пункты положения о закупках (Contract) пока не делаю, потому что еще с этим не разобрались

        self.dict_data_f1_db = {}
        self.dict_data_f1_db["pn_ext_ident"] = [pn_ext_ident]
        self.dict_data_f1_db["ID_OKVED2"] = [ID_OKVED]
        self.dict_data_f1_db["ID_OKPD2"] = [ID_OKPD]
        self.dict_data_f1_db["Object"] = [Object]
        self.dict_data_f1_db["DeliveryPlace"] = [DeliveryPlace]
        self.dict_data_f1_db["StartPrice"] = [StartPrice]
        self.dict_data_f1_db["DataStart"] = [DataStart]
        self.dict_data_f1_db["DataStop"] = [DataStop]
        self.dict_data_f1_db["ID_PurchaseWay"] = [ID_PurchaseWay]
        self.dict_data_f1_db["ID_PurchaseCondition"] = [ID_PurchaseCondition]
        # self.dict_data_f1_db["ID_Breakdown"] пока нету
        self.dict_data_f1_db["ID_FinancialSources"] = [ID_FinancialSources]
        # self.dict_data_f1_db["Contract"] пока нету

        conn.close()
        return True

    def check_exist(self, j):
        self.j = j
        for key in self.dict_data.keys():
            if len(self.dict_data[key][0]) == 0:
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля в строчке %s!!!" % (int(self.j / 2)))
                mess.exec()
                return False
            elif self.dict_data[key][0] == "**.**.****":
                mess = QtWidgets.QMessageBox(self)
                mess.setText("Заполните все поля в строчке %s!!!" % (int(self.j / 2)))
                mess.exec()
                return False

    def insert_in_db(self):
        conn = psycopg2.connect(dbname='postgres', user='postgres',
                                password='300771Den', host='localhost')
        cursor = conn.cursor()

        query = "call public.save_plan_position(\n" \
                "\tpid_plan => 1,\n" \
                "\tpv_position_status => 'A',\n" \
                "\tpid_okved2 => {},\n" \
                "\tpid_okpd2 => {},\n" \
                "\tpv_purchase_object => '{}',\n" \
                "\tpv_delivery_place => '{}',\n" \
                "\tpf_start_max_price => {},\n" \
                "\tpdt_purchase_start => '{}',\n" \
                "\tpdt_purchase_stop => '{}',\n" \
                "\tpid_purchase_way => {},\n" \
                "\tpid_purchase_condition => {},\n" \
                "\tpid_payment_breakdown => null,\n" \
                "\tpid_financing_source => {},\n" \
                "\tpid_user => 1,\n" \
                "\tpid_region => 1,\n" \
                "\tpn_ext_ident => {},\n" \
                "\tpaction => 1);".format(self.dict_data_f1_db["ID_OKVED2"][0][0],
                                          self.dict_data_f1_db["ID_OKPD2"][0][0],
                                          self.dict_data_f1_db["Object"][0], self.dict_data_f1_db["DeliveryPlace"][0],
                                          self.dict_data_f1_db["StartPrice"][0], self.dict_data_f1_db["DataStart"][0],
                                          self.dict_data_f1_db["DataStop"][0],
                                          self.dict_data_f1_db["ID_PurchaseWay"][0][0],
                                          self.dict_data_f1_db["ID_PurchaseCondition"][0][0],
                                          self.dict_data_f1_db["ID_FinancialSources"][0][0],
                                          self.dict_data_f1_db["pn_ext_ident"][0])

        cursor.execute(query)
        conn.commit()
        conn.close()
        return None



    def openFileNameDialog(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        fileName, _ = QtWidgets.QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                            "All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            res = self.read_xls(fileName)
            self.dialog = Second(res, self.task)
            self.dialog.setAttribute(QtCore.Qt.WA_DeleteOnClose)
            self.dialog.show()

    def read_xls(self, fileName):
        wb = load_workbook(fileName)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        table = np.zeros(13).reshape(1, 13)
        array = np.zeros(13).reshape(1, 13)
        i = 19
        cycle = True
        count = 0
        while cycle == True:
            check_array = [[str(type(cell.value)) for cell in col] for col in sheet['B' + str(i):'N' + str(i)]]
            true_array = [[cell.value for cell in col] for col in sheet['B' + str(i):'N' + str(i)]]
            unique, counts = np.unique(check_array, return_counts=True)
            print(unique, counts)
            for j in range(unique[np.where(counts == max(counts))].shape[0]):
                if unique[np.where(counts == max(counts))][j] != "<class 'NoneType'>" and counts[
                    np.where(counts == max(counts))[0][j]] >= 6:
                    array = np.append(array, true_array, axis=0)
                elif counts[np.where(counts == max(counts))[0][j]] == 13:
                    count += 1
                    if count > 5:
                        cycle = False
            i += 1
        if array[1][0] == '1':
            return(array[2:])
        else:
            return(array[1:])

# атрибут модуля _nате_ будет содержать значение "_main_" только в случае заnуска модуля как главной nрограммы.
if __name__ == '__main__':
    import \
        sys  # подключение модуля sys, из которого нам потребуется список параметров, переданных в командной строке (argv)

    app = QtWidgets.QApplication(sys.argv)  # создает объект приложения с помощрю класса QApplication
    window = LoginPage()
    window.show()
    sys.exit(app.exec_())  # функция, позволяющая завершить выполнение программы
